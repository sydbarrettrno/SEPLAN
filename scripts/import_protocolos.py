import argparse
import csv
import json
import os
import sys
import unicodedata
from datetime import date, datetime, time
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional

from sqlmodel import Session, SQLModel, create_engine, select

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from models import ProtocolRecord  # noqa: E402
from services_protocolos import build_texto_busca  # noqa: E402


FIELD_ALIASES: Mapping[str, Iterable[str]] = {
    "protocolo": ("protocolo", "numero", "n_protocolo", "numero_protocolo"),
    "ano": ("ano", "exercicio"),
    "numero_ano": ("numero_ano", "numero/ano", "protocolo_ano", "protocolo/ano"),
    "requerente": ("requerente", "interessado", "solicitante"),
    "abertura_data": ("abertura_data", "data_abertura", "abertura"),
    "obs_abertura": ("obs_abertura", "observacao_abertura", "observacoes_abertura"),
    "ultimo_tramite_data": ("ultimo_tramite_data", "data_ultimo_tramite", "ultimo_tramite"),
    "situacao": ("situacao", "status", "situacao_atual"),
    "subassunto": ("subassunto", "assunto", "sub_assunto"),
    "ultimo_tramite_obs": ("ultimo_tramite_obs", "observacao_ultimo_tramite", "obs_ultimo_tramite"),
    "responsavel": ("responsavel", "responsavel_atual", "setor_responsavel"),
    "tipo_processo_normalizado": ("tipo_processo_normalizado", "tipo_processo", "tipo"),
    "status_semantico_normalizado": ("status_semantico_normalizado", "status_semantico"),
    "confiabilidade_extracao": ("confiabilidade_extracao", "confiabilidade"),
    "status_revisao": ("status_revisao", "revisao"),
    "texto_busca": ("texto_busca", "texto"),
}


def main() -> int:
    parser = argparse.ArgumentParser(description="Importa protocolos normalizados de CSV ou XLSX.")
    parser.add_argument("file_path", help="Caminho do arquivo CSV ou XLSX.")
    parser.add_argument(
        "--database-url",
        default=os.getenv("DATABASE_URL", "sqlite:///./app.db"),
        help="URL do banco. Padrao: DATABASE_URL ou sqlite:///./app.db.",
    )
    args = parser.parse_args()

    file_path = Path(args.file_path)
    if not file_path.exists():
        print(json.dumps({"status": "error", "message": "Arquivo nao encontrado", "file": str(file_path)}))
        return 1

    engine = create_engine(
        args.database_url,
        connect_args={"check_same_thread": False} if args.database_url.startswith("sqlite") else {},
    )
    SQLModel.metadata.create_all(engine)

    rows = list(read_rows(file_path))
    imported = 0
    updated = 0
    skipped = 0

    with Session(engine) as session:
        for raw_row in rows:
            data = normalize_row(raw_row)
            record = build_record(data)
            if record is None:
                skipped += 1
                continue

            existing = find_existing(session, record)
            if existing:
                update_record(existing, record)
                updated += 1
            else:
                session.add(record)
                imported += 1

        session.commit()

    print(
        json.dumps(
            {
                "status": "ok",
                "file": str(file_path),
                "total_rows": len(rows),
                "imported": imported,
                "updated": updated,
                "skipped": skipped,
            },
            ensure_ascii=False,
        )
    )
    return 0


def read_rows(file_path: Path) -> Iterable[Dict[str, object]]:
    suffix = file_path.suffix.casefold()
    if suffix == ".csv":
        return read_csv_rows(file_path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(file_path)

    raise SystemExit(f"Formato nao suportado: {suffix}. Use CSV ou XLSX.")


def read_csv_rows(file_path: Path) -> List[Dict[str, object]]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with file_path.open("r", encoding=encoding, newline="") as handle:
                return list(csv.DictReader(handle))
        except UnicodeDecodeError:
            continue

    raise SystemExit("Nao foi possivel ler o CSV com codificacao conhecida.")


def read_xlsx_rows(file_path: Path) -> List[Dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Para importar XLSX, instale a dependencia openpyxl.") from exc

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value or "").strip() for value in rows[0]]
    parsed_rows: List[Dict[str, object]] = []
    for values in rows[1:]:
        parsed_rows.append({headers[index]: value for index, value in enumerate(values) if index < len(headers)})

    return parsed_rows


def normalize_row(row: Mapping[str, object]) -> Dict[str, object]:
    normalized_headers = {canonicalize(header): value for header, value in row.items()}
    data: Dict[str, object] = {}

    for target_field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            value = normalized_headers.get(canonicalize(alias))
            if value is not None and str(value).strip():
                data[target_field] = value
                break

    return data


def canonicalize(value: object) -> str:
    text = str(value or "").casefold().strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = "".join(char if char.isalnum() else "_" for char in text)
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_")


def build_record(data: Mapping[str, object]) -> Optional[ProtocolRecord]:
    protocolo = clean_text(data.get("protocolo")) or clean_text(data.get("numero_ano"))
    if not protocolo:
        return None

    ano = parse_int(data.get("ano")) or infer_year(data.get("numero_ano"))
    normalized_data = {
        "protocolo": protocolo,
        "ano": ano,
        "numero_ano": clean_text(data.get("numero_ano")),
        "requerente": clean_text(data.get("requerente")),
        "abertura_data": parse_date(data.get("abertura_data")),
        "obs_abertura": clean_text(data.get("obs_abertura")),
        "ultimo_tramite_data": parse_datetime(data.get("ultimo_tramite_data")),
        "situacao": clean_text(data.get("situacao")),
        "subassunto": clean_text(data.get("subassunto")),
        "ultimo_tramite_obs": clean_text(data.get("ultimo_tramite_obs")),
        "responsavel": clean_text(data.get("responsavel")),
        "tipo_processo_normalizado": clean_text(data.get("tipo_processo_normalizado")),
        "status_semantico_normalizado": clean_text(data.get("status_semantico_normalizado")),
        "confiabilidade_extracao": parse_float(data.get("confiabilidade_extracao")),
        "status_revisao": clean_text(data.get("status_revisao")) or "pendente",
        "texto_busca": clean_text(data.get("texto_busca")),
    }
    if not normalized_data["numero_ano"] and ano:
        normalized_data["numero_ano"] = f"{protocolo}/{ano}"

    if not normalized_data["texto_busca"]:
        normalized_data["texto_busca"] = build_texto_busca(normalized_data)

    return ProtocolRecord(**normalized_data)


def find_existing(session: Session, record: ProtocolRecord) -> Optional[ProtocolRecord]:
    if record.ano is not None:
        existing = session.exec(
            select(ProtocolRecord).where(
                ProtocolRecord.protocolo == record.protocolo,
                ProtocolRecord.ano == record.ano,
            )
        ).first()
        if existing:
            return existing

    if record.numero_ano:
        return session.exec(
            select(ProtocolRecord).where(ProtocolRecord.numero_ano == record.numero_ano)
        ).first()

    return session.exec(
        select(ProtocolRecord).where(ProtocolRecord.protocolo == record.protocolo)
    ).first()


def update_record(existing: ProtocolRecord, incoming: ProtocolRecord) -> None:
    fields = (
        "protocolo",
        "ano",
        "numero_ano",
        "requerente",
        "abertura_data",
        "obs_abertura",
        "ultimo_tramite_data",
        "situacao",
        "subassunto",
        "ultimo_tramite_obs",
        "responsavel",
        "tipo_processo_normalizado",
        "status_semantico_normalizado",
        "confiabilidade_extracao",
        "status_revisao",
        "texto_busca",
    )
    for field_name in fields:
        setattr(existing, field_name, getattr(incoming, field_name))
    existing.updated_at = datetime.utcnow()


def clean_text(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text or None


def parse_int(value: object) -> Optional[int]:
    text = clean_text(value)
    if not text:
        return None
    try:
        return int(float(text.replace(",", ".")))
    except ValueError:
        return None


def parse_float(value: object) -> Optional[float]:
    text = clean_text(value)
    if not text:
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def parse_date(value: object) -> Optional[date]:
    parsed = parse_datetime(value)
    return parsed.date() if parsed else None


def parse_datetime(value: object) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)

    text = clean_text(value)
    if not text:
        return None

    for date_format in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y",
        "%d-%m-%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            continue

    return None


def infer_year(value: object) -> Optional[int]:
    text = clean_text(value)
    if not text:
        return None
    parts = text.replace("-", "/").split("/")
    for part in reversed(parts):
        year = parse_int(part)
        if year and 1900 <= year <= 2100:
            return year
    return None


if __name__ == "__main__":
    raise SystemExit(main())
